from openpyxl import Workbook
import datetime

wb = Workbook()
ws = wb.active
ws.title = "Expense Tracker"

# Row Titles
ws['A1'] = 'Variable Expenses'
ws['A2'] = 'Rent'
ws['A3'] = 'Gas'
ws['A4'] = 'Groceries'
ws['A5'] = 'Restaurants'
ws['A6'] = 'Student Loan'
ws['A6'] = 'Savings'
ws['A7'] = 'Recreation'

# Formatting
ws.column_dimensions['A'].width = 20
    
# Dates
# ws['B2'] = datetime.datetime(2019, 1, 9).strftime("%m/%d/%y")

# Column Titles
ws['B1'] = 'Budgeted'
ws['C1'] = 'Spent'
ws['D1'] = 'Remaining'

for row in range(2, 8):
     for col in range(2, 5):
        _ = ws.cell(column=col, row=row, value="{0}".format("0"))


for row in range(2, 8):
     for col in range(4, 5):
        _ = ws.cell(column=col, row=row, value="{0}".format("=(B2:B7 - C2:C7)"))

# Budgeted amounts
ws['B2'] = 900
ws['B3'] = 200
ws['B4'] = 200

# Spent amounts

# Spent amounts


wb.save("expense_tracker.xlsx")
