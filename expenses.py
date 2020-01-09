from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import datetime

wb = Workbook()
ws = wb.active
ws.title = "Expense Tracker"

# Row Titles
ws['A1'] = 'Variable Expenses'
ws['A2'] = 'Rent'
ws['A3'] = 'Gas'
ws['A4'] = 'Groceries'
ws['A5'] = 'Restaurants'
ws['A6'] = 'Student Loan'
ws['A6'] = 'Savings'
ws['A7'] = 'Recreation'

for row in range(2, 8):
     for col in range(3, 6):
        _ = ws.cell(column=col, row=row, value="{0}".format(0))

ws['B2'] = datetime.datetime(2019, 1, 9).strftime("%m/%d/%y")

# Column Titles
ws['B1'] = 'Date'
ws['C1'] = 'Budgeted'
ws['D1'] = 'Spent'
ws['E1'] = 'Remaining'


for row in range(2, 8):
     for col in range(5, 6):
        _ = ws.cell(column=col, row=row, value="{0}".format("=(C2:C7 - D2:D7)"))

# Budgeted amounts
ws['C2'] = 900
ws['C3'] = 200
ws['C4'] = 200

# Spent amounts
ws['D3'] = 168.71


wb.save("expense_tracker.xlsx")
