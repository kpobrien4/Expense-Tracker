import click
import pandas as pd
from openpyxl import Workbook, load_workbook
import os
import datetime


wb = load_workbook(filename = "expense_tracker.xlsx")
ws = wb.active

os.system("python3 spreadsheet.py")
expenses = pd.read_excel("expense_tracker.xlsx")
   
@click.command()
@click.option('--category', prompt='Category of expense', help='Category money was spent in.')
@click.option('--expense', prompt='Amount spent', default=0.0, help='Amount spent.')


def categoryValid(expense, category):
    """Simple program that tracks EXPENSES in various CATEGORIES."""
    if category in expenses["Variable Expenses"].values:
        if category == "Rent":
            spentCell = 'C2'
            budgetCell = 'B2'
            spent(expense, category, spentCell, budgetCell)
        elif category == "Gas":
            spentCell = 'C3'
            budgetCell = 'B3'
            spent(expense, category, spentCell, budgetCell)
        elif category == "Groceries":
            spentCell = 'C4'
            budgetCell = 'B4'
            spent(expense, category, spentCell, budgetCell)
        elif category == "Restaurants":
            spentCell = 'C5'
            budgetCell = 'B5'
            spent(expense, category, spentCell, budgetCell)
        elif category == "Savings":
            spentCell = 'C6'
            budgetCell = 'B6'
            spent(expense, category, spentCell, budgetCell)
        elif category == "Recreation":
            spentCell = 'C7'
            budgetCell = 'B7'
            spent(expense, category, spentCell, budgetCell)
        else:
            print("That is not a valid category")
    else:
        print("That is not a valid category")

def spent(expense, category, spentCell, budgetCell):
        click.echo('Spent ${:,.2f} on %s'.format(expense) % ( category))
        addSpent(category, expense, spentCell)
        balance(category, spentCell, budgetCell)
    
def addSpent(cat, amt, spentCell):
    if amt > 0:
        ws[spentCell] = float(ws[spentCell].value) + float(amt)
        wb.save("expense_tracker.xlsx")
    else: 
        print("Please refrain from spending negative money")

def balance(cat, spentCell, budgetCell):
    chosenCategory = cat

    totalAmt = float(ws[budgetCell].value) - float(ws[spentCell].value)
    if totalAmt < 0:
        print("You are $%.2f over budget for %s" % (totalAmt, chosenCategory))
    elif totalAmt > 0:
        print("You have $%.2f remaining for %s" % (totalAmt, chosenCategory))
    else:
        print("You are at budget for %s" % chosenCategory)
    printTotal()

def printTotal():
    expensesAfter = pd.read_excel("expense_tracker.xlsx")
    totalSpent = expensesAfter["Spent ($)"].sum()
    print("You have spent a total of ${:,.2f} this month.".format(totalSpent))

if __name__ == '__main__':
    categoryValid()



